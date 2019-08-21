import os

import openpyxl

from django.contrib.auth import get_user_model
from django.contrib.auth.management import get_default_username
from django.contrib.auth.password_validation import validate_password
from django.core import exceptions
from django.core.management.base import BaseCommand, CommandError
from django.db import (
    IntegrityError, DEFAULT_DB_ALIAS
)
from django.utils.text import capfirst

from core.models import (
    User, Batch, Section
)


class Command(BaseCommand):
    help = 'Used to create the voters based from an XLSX file.'
    requires_migrations_checks = True
    stealth_options = ('stdin',)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            help='Specifies the XLSX file the voter information are stored in.'
        )

        parser.add_argument(
            '--outputfile',
            help='Specifies the XLSX file to put the voter credentials in.'
        )

    def execute(self, *args, **options):
        self.stdin = options.get('stdin', sys.stdin)
        return super().execute(*args, **options)

    def handle(self, *args, **options):
        voter_xlsx_filepath = os.path.abspath(options['file'])
        voter_data_wb = openpyxl.load_workbook(voter_xlsx_filepath)
        voter_credentials_wb = openpyxl.Workbook()

        # File Format
        # A - Last Name
        # B - First Name
        # C - Batch
        # D - Section
        row_counter = 0
        for row in voter_data_wb.iter_rows(min_row=2):
            voter_data = [ col.value for col in row ]

            first_name = voter_data[1]
            last_name = voter_data[0]
            username = '{}{}'.format(
                first_name.split()[0].lower(),
                last_name.lower()
            )
            batch_year = voter_data[2]
            section_name = voter_data[3]

            try:
                batch = Batch.objects.get(year=batch_year)
            except Batch.DoesNotExist:
                batch = Batch.objects.create(year=batch_year)

            try:
                section = Section.objects.get(section_name=section_name)
            except Section.DoesNotExist:
                section = Section.objects.create(section_name=section_name)

            try:
                user = User.objects.create(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    batch=batch,
                    section=section
                )
            except IntegrityError:
                # Likely caused by duplicate username.
                username = '{}{}'.format(username, row_counter)
            finally:
                user = User.objects.create(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    batch=batch,
                    section=section
                )

            user.set_password(username)
            user.save()

            # Save information to the voter credentials file.
            voter_credentials_wb['A{}'.format(row_counter)] = last_name
            voter_credentials_wb['B{}'.format(row_counter)] = first_name
            voter_credentials_wb['C{}'.format(row_counter)] = username

            row_counter += 1

        voter_credentials_filepath = os.path.abspath(options['outputfile'])
        voter_credentials_wb.save(voter_credentials_filepath)
